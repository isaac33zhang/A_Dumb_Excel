# -*- coding: utf-8 -*-
import os
import openpyxl
from lib.pangpang_helper import PangPangHelper


DEPARTMENT_PATH = "/Users/fanzhaf/Documents/Personal/PyCharm/PycharmProjects/excel/actual_departments.xlsx"
SALARY_PATH = "/Users/fanzhaf/Documents/Personal/PyCharm/PycharmProjects/excel/salary.xlsx"
DESTINATION_PATH = "/Users/fanzhaf/Desktop/"


def test_read_data_from_departments():
    helper = PangPangHelper(DEPARTMENT_PATH, SALARY_PATH, DESTINATION_PATH)
    data = helper.read_data_from_departments()
    key = data.keys()[0]
    people = 0
    for key in data:
        people = people + len(data[key])
    assert people == 187
    assert len(data[key]) == 4
    assert len(data.keys()) == 36


def test_read_data_from_salary():
    helper = PangPangHelper(DEPARTMENT_PATH, SALARY_PATH, DESTINATION_PATH)
    data = helper.read_data_from_salary()
    key = data.keys()[0]
    assert len(data[key]) == 29
    assert len(data.keys()) == 141
    assert helper.item_names
    assert len(helper.item_names) == 29


def test_generate_new_forms():
    helper = PangPangHelper(DEPARTMENT_PATH, SALARY_PATH, DESTINATION_PATH)
    department_data = helper.read_data_from_departments()
    salary_data = helper.read_data_from_salary()
    helper.generate_new_forms(department_data, salary_data)
    name = department_data.keys()[0]
    path = DESTINATION_PATH + name + ".xlsx"
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    assert worksheet.max_row == 6
    assert worksheet.max_column == 29
    for name in department_data.keys():
        print(name)
        path = DESTINATION_PATH + name + ".xlsx"
        os.remove(path)

