import openpyxl
from openpyxl.utils import get_column_letter


class ExcelObject(object):
    def __init__(self, file_path):
        self.workbook = openpyxl.load_workbook(file_path)

    @staticmethod
    def create_file(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    def get_sheet_objects(self):
        sheet_names = self.workbook.get_sheet_names()
        return [self.workbook.get_sheet_by_name(name) for name in sheet_names]

    def get_sheet(self, sheet_name=None):
        if sheet_name:
            return self.workbook.get_sheet_by_name(sheet_name)
        else:
            return self.get_sheet_objects()[0]

    def get_cell(self, row=1, column=1, sheet_name=None):
        sheet = self.get_sheet(sheet_name)
        cell = sheet.cell(row=row, column=column)
        return cell.value

    def get_row_count(self, sheet_name=None):
        sheet = self.get_sheet(sheet_name)
        return sheet.max_row

    def get_column_count(self, sheet_name=None):
        sheet = self.get_sheet(sheet_name)
        return sheet.max_column

    def get_row_list(self, row):
        sheet = self.get_sheet()
        max_column = self.get_column_count()
        column_letter = get_column_letter(max_column)
        start = 'A' + str(row)
        end = column_letter + str(row)
        return sheet[start: end][0]



