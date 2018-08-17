# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
from excel_utils import ExcelObject


class PangPangHelper(object):
    def __init__(self, actual_departments, salary, destination):
        self.departments_wb = ExcelObject(actual_departments)
        self.salary_wb = ExcelObject(salary)
        self.dest_path = destination
        self.item_names = None

    def run(self):
        department_data = self.read_data_from_departments()
        salary_data = self.read_data_from_salary()
        self.generate_new_forms(department_data, salary_data)

    def read_data_from_departments(self):
        data = {}
        row_number = self.departments_wb.get_row_count()
        print("[Departments] Row count is {}".format(row_number))
        print("[Departments] Reading rows")
        for row in range(2, row_number+1):
            row_list = self.departments_wb.get_row_list(row)
            if not row_list[0].value:
                continue
            if row_list[6].value in data:
                data[row_list[6].value].append(row_list[1].value)
            else:
                data[row_list[6].value] = [row_list[1].value]
        print("[Departments] Read {} people for {} departments".format(row_number-1, len(data.keys())))
        return data

    def read_data_from_salary(self):
        data = {}
        row_number = self.salary_wb.get_row_count()
        print("[Salary] Row count is {}".format(row_number))
        print("[Salary] Reading rows")
        for row in range(1, row_number+1):
            if self.salary_wb.get_cell(row=row, column=1):
                if self.salary_wb.get_cell(row=row, column=1).encode('utf-8') == "统发银行名称":
                    self.item_names = self.salary_wb.get_row_list(row)
                    break
        for row in range(1, row_number+1):
            if self.salary_wb.get_cell(row=row, column=1):
                if self.salary_wb.get_cell(row=row, column=1).encode('utf-8') == "工商银行":
                    row_list = self.salary_wb.get_row_list(row)
                    data[row_list[2].value] = row_list
        print("[Salary] Read {} people from salary form".format(len(data.keys())))
        return data

    def _create_single_file(self, name, data):
        print("[Generation] Creating file for " + name)
        workbook = openpyxl.Workbook()
        path = self.dest_path + name + ".xlsx"
        print("[Generation] Creating file at" + path)
        workbook.save(path)
        worksheet = workbook.active
        print("[Generation] Writing {} items each person for {} people".format(len(data[0]), len(data)))
        for row in range(len(data)):
            for column in range(len(data[0])):
                column_letter = get_column_letter(column+1)
                name_index = row * 2 + 1
                data_index = row * 2 + 2
                cell_index = column_letter + str(name_index)
                worksheet[cell_index] = self.item_names[column].value
                cell_index = column_letter + str(data_index)
                worksheet[cell_index] = data[row][column].value
        workbook.save(path)

    def generate_new_forms(self, department_data, salary_data):
        for department in department_data:
            data = []
            for name in department_data[department]:
                try:
                    data.append(salary_data[name])
                except KeyError as e:
                    print("[Generation] " + name + " is not in salary file")
            self._create_single_file(department, data)
        print("[Generation] Finished generating files")


